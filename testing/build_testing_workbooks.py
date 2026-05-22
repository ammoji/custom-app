"""
Builds two Excel deliverables for the Kirana Mart family-testing team:

1. kirana-mart-features.xlsx
   A catalog of every user-facing feature shipped so far.
   Filterable by role, category, status. To be updated as new PRs ship.

2. kirana-mart-test-cases.xlsx
   Detailed test cases across positive / negative / edge case types.
   Includes blank columns testers fill in: Actual Result, Status,
   Severity, Tested By, Tested Date, Comments.

Both files use Arial throughout, frozen header rows, color-coded
headers, and data-validation dropdowns where Google Sheets supports
them after upload.

USAGE (on Windows PowerShell):
    pip install openpyxl
    python build_testing_workbooks.py

Both .xlsx files are written next to this script.
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
import os

# Output destination = same folder as this script.
OUTPUT_DIR = os.path.dirname(os.path.abspath(__file__))

# ----- Shared style constants -----
PRIMARY_GREEN = '0E7C3A'
LIGHT_GREEN = 'DCFCE7'
LIGHT_YELLOW = 'FEF3C7'
HEADER_FILL = PatternFill('solid', start_color=PRIMARY_GREEN)
ZEBRA_FILL = PatternFill('solid', start_color='F9FAFB')

HEADER_FONT = Font(name='Arial', size=11, bold=True, color='FFFFFF')
BODY_FONT = Font(name='Arial', size=10)
THIN = Side(border_style='thin', color='D1D5DB')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP_TOP = Alignment(wrap_text=True, vertical='top', horizontal='left')


# ============================================================================
# FEATURES INVENTORY
# (ID, Category, Role, Feature Name, Description, How To Access,
#  Status, PR Reference, Notes)
# ============================================================================
FEATURES = [
    # ---- Authentication & Profile ----
    ('FE-001', 'Authentication', 'All Users',
     'Phone OTP login',
     'Sign in or sign up using your Indian mobile number and a 6-digit OTP sent via SMS.',
     'App opens to Sign in screen; enter 10-digit number, tap Send OTP.',
     'Live', '-', 'Indian +91 numbers only. Test phones bypass real SMS.'),
    ('FE-002', 'Authentication', 'All Users',
     'Resend OTP with cooldown',
     "If SMS doesn't arrive, tap 'Didn't get the code? Resend OTP'. 30-second cooldown prevents spam.",
     'OTP entry screen, link below the Verify button.',
     'Live', 'PR 10', 'Cooldown countdown displayed live.'),
    ('FE-003', 'Authentication', 'All Users',
     'Profile setup with mandatory name',
     'First-time users must enter their full name before continuing. Phone number is auto-populated and not editable.',
     'After OTP verification, navigates to Profile screen.',
     'Live', 'PR 10', 'Phone is identity anchor and immutable.'),
    ('FE-004', 'Authentication', 'All Users',
     'Sign out',
     'Sign out clears all local app state including cart and persisted session.',
     'Profile screen, scroll to bottom, tap Sign Out.',
     'Live', '-', 'Local cart and any saved-for-later data is wiped.'),
    ('FE-005', 'Authentication', 'All Users',
     'Multi-role support',
     'A single phone number can be Customer, Shop Owner, Delivery Partner, and Admin simultaneously.',
     'Admin assigns roles; user sees additional dashboards under Your Roles on Home.',
     'Live', '-', 'Custom Firebase claims drive role-based access.'),
    ('FE-006', 'Authentication', 'Customer',
     'Address management',
     'Save multiple delivery addresses (Home, Work, etc.). Set a default address.',
     'Profile screen, Saved Addresses section.',
     'Live', '-', 'Used at checkout.'),

    # ---- Customer: Browse & Search ----
    ('FE-010', 'Browse', 'Customer',
     'Browse shops near you',
     "List of active shops. Distance filter is currently OFF for testing so testers across India can see each other's shops.",
     'Home screen → Browse shops near me, or Search → Browse all.',
     'Live', 'PR 10', 'Production launch will re-enable 1km radius.'),
    ('FE-011', 'Browse', 'Customer',
     'Shop detail screen',
     "Opens a shop's full menu organized by category. Shows shop name, address, delivery fee, minimum order.",
     'Tap any shop card from Browse, Search, or HomeScreen.',
     'Live', '-', 'Menu items fetched from per-shop menu collection.'),
    ('FE-012', 'Browse', 'Customer',
     'Search products across shops',
     'Type a product name (e.g. atta, milk, soap) to find which shops carry it.',
     'Home screen → Search bar at top, or bottom-tab Search.',
     'Live', 'PR 4', 'Cross-shop search returns menu items + shop names.'),
    ('FE-013', 'Browse', 'Customer',
     'Category filter',
     'Filter shops or search results by category (Atta/Rice/Dal, Beverages, Personal Care, etc.).',
     'Home screen → tap a category chip, or Search → tap category.',
     'Live', '-', 'Categories defined in constants/categories.ts.'),

    # ---- Customer: Cart ----
    ('FE-020', 'Cart', 'Customer',
     'Add to cart',
     'Tap + on any menu item to add. Quantity increments. Cart icon shows running total.',
     'Shop Detail screen or Search results.',
     'Live', '-', 'Cart persists across app restarts.'),
    ('FE-021', 'Cart', 'Customer',
     'Adjust quantity',
     'Increment or decrement item quantity. Decrementing to 0 removes the item.',
     'Cart screen, or directly from Shop Detail screen.',
     'Live', '-', 'Cart price updates live.'),
    ('FE-022', 'Cart', 'Customer',
     'Multi-shop cart blocker',
     "Cart can only hold items from one shop at a time. Adding from a second shop prompts 'Replace cart?'.",
     'Try adding from a different shop after building a cart.',
     'Live', 'PR 4', 'Server validates cart-line shop consistency at place-order.'),
    ('FE-023', 'Cart', 'Customer',
     'Cart persistence',
     'Cart contents survive app close, reopen, even sign-out / sign-in within the same install.',
     'Build cart, force-close app, reopen — items still there.',
     'Live', '-', 'Zustand persist via AsyncStorage. Versioned (cart-v2).'),

    # ---- Customer: Checkout & Payment ----
    ('FE-030', 'Checkout', 'Customer',
     'Address selection at checkout',
     'Pick from saved addresses, or add a new one inline.',
     'Cart → Proceed to checkout → Delivery address section.',
     'Live', '-', 'Defaults to most recent address.'),
    ('FE-031', 'Checkout', 'Customer',
     'Choose payment method',
     'Cash on Delivery (COD) or Online (Razorpay). Online supports cards, UPI, netbanking.',
     'Checkout screen, Payment section.',
     'Live', '-', 'Razorpay uses test keys on dev project.'),
    ('FE-032', 'Checkout', 'Customer',
     'Online payment via Razorpay',
     "Razorpay's payment overlay opens. Enter card / UPI / netbanking details. Webhook confirms on success.",
     "Checkout → Online (Razorpay) → Place Order → Razorpay overlay opens.",
     'Live', 'PR 2', 'Indian test cards only (e.g. 4386 2894 0766 0153).'),
    ('FE-033', 'Checkout', 'Customer',
     'Payment failure handling',
     "If payment fails or is dismissed, order stays in 'Payment pending' state. Customer can retry payment or cancel.",
     'Trigger by dismissing Razorpay overlay or using failure test card.',
     'Live', 'PR 2', 'Auto-cancellation after 24h via cleanup cron.'),
    ('FE-034', 'Checkout', 'Customer',
     'Minimum order enforcement',
     "Cart subtotal must meet the shop's minimum order before checkout is allowed.",
     'Cart with subtotal below minimum shows "Add X more to proceed" warning.',
     'Live', 'PR 5', 'Server validates again at placeOrder.'),

    # ---- Customer: Orders ----
    ('FE-040', 'Customer Orders', 'Customer',
     'Active orders rail on Home',
     "Live in-flight orders appear as cards at the top of Home with status chip and ETA. Tap to view detail.",
     'HomeScreen, top section, only visible when you have active orders.',
     'Live', 'PR 15', 'Reuses listMine fetch; auto-hides when no active orders.'),
    ('FE-041', 'Customer Orders', 'Customer',
     'Order Again rail on Home',
     "After delivered orders accumulate, the shops you order from most appear as 'Order again from X' cards.",
     'HomeScreen, below Active Orders rail.',
     'Live', 'PR 14', 'Top 3 by frequency; auto-hides for new users.'),
    ('FE-042', 'Customer Orders', 'Customer',
     'Reorder button on past orders',
     "Tap Reorder on any delivered or cancelled order. Modal shows item availability + current prices. Confirm to fill cart.",
     'My Orders tab → tap Reorder on an eligible card.',
     'Live', 'PR 13', "Replaces cart entirely - Zomato/Swiggy convention."),
    ('FE-043', 'Customer Orders', 'Customer',
     'View order detail',
     'Full order info: items with prices, delivery address, payment method/status, current order status, full timeline.',
     'My Orders tab → tap any order card; or Home active rail → tap card.',
     'Live', '-', 'Polling cadence ~5s on native.'),
    ('FE-044', 'Customer Orders', 'Customer',
     'Cancel pending unpaid order',
     'If order is still pending and unpaid (e.g. Razorpay was dismissed), customer can cancel and place new order.',
     'Order detail screen for a pending payment-failed order.',
     'Live', '-', 'COD orders also cancellable while still pending.'),
    ('FE-045', 'Customer Orders', 'Customer',
     'Cancel paid order within 2-min window',
     'For 2 minutes after a paid order is placed, customer can self-cancel with automatic refund.',
     'Order detail of a paid order < 2 minutes old. Shows countdown.',
     'Live', 'PR 7', 'After 2 min, customer must contact support.'),
    ('FE-046', 'Customer Orders', 'Customer',
     'Refund status visibility',
     'After a cancelled paid order, customer sees Refund pending / Refunded / Refund failed states on order detail.',
     'Order detail of a cancelled paid order.',
     'Live', 'PR 7 hotfix', 'Razorpay refund usually completes within seconds.'),
    ('FE-047', 'Customer Orders', 'Customer',
     'Customer-facing status copy',
     "Customers see familiar 'Out for delivery' phrasing even when internal status is 'ready_for_pickup'.",
     'Order detail and active orders rail when shop marks order Ready.',
     'Live', 'PR 12', 'Shopkeepers and delivery partners see internal labels.'),

    # ---- Shop Owner: Registration ----
    ('FE-050', 'Shop Registration', 'Shop Owner',
     'Register your shop',
     'Submit shop name, address, contact, FSSAI/GSTIN (optional), pickup location. Goes to admin for approval.',
     'Home → Become a Shop Owner tile (if no shop) → fill form.',
     'Live', '-', 'Pending shop status until admin approves.'),
    ('FE-051', 'Shop Registration', 'Shop Owner',
     'Wait for approval',
     'After submitting, a Waiting for Approval screen shows status + rejection reason if any.',
     'Home → Awaiting approval tile.',
     'Live', '-', 'Rejected shops can re-submit with prefill.'),

    # ---- Shop Owner: Dashboard & Orders ----
    ('FE-060', 'Shop Orders', 'Shop Owner',
     'Shop dashboard with active orders',
     'List of all active orders (pending, accepted, preparing, ready_for_pickup) for your shop only.',
     'Home → Shop dashboard tile.',
     'Live', '-', 'Polls every ~10s. Scoped server-side to your shopId.'),
    ('FE-061', 'Shop Orders', 'Shop Owner',
     'Pull to refresh',
     'Pull down on the order list to manually refresh.',
     'Shop dashboard, pull down.',
     'Live', 'PR 7', 'Same on AdminOrders.'),
    ('FE-062', 'Shop Orders', 'Shop Owner',
     "Today's stats card",
     "Above the order list, see today's count and revenue at a glance.",
     'Top of shop dashboard.',
     'Live', '-', 'Counts orders created today, in any status.'),
    ('FE-063', 'Shop Orders', 'Shop Owner',
     'Accept order with mandatory ETA',
     "Tap Accept on a pending order. A modal opens asking 'Ready in [X] minutes?'. Enter 1-240 minutes, submit.",
     'Shop Order detail → Accept button.',
     'Live', 'PR 12', 'Server rejects accept without ETA.'),
    ('FE-064', 'Shop Orders', 'Shop Owner',
     'Update ETA when starting to prepare',
     "Tap Start Preparing. Modal opens with current ETA prefilled. Adjust if running late.",
     'Shop Order detail (accepted state) → Start Preparing button.',
     'Live', 'PR 12', 'Admin dashboard shows "updated from" if changed.'),
    ('FE-065', 'Shop Orders', 'Shop Owner',
     'Mark Ready for Pickup',
     "Final action shop takes - signals delivery partners the order is physically ready.",
     'Shop Order detail (preparing state) → Ready for Pickup button.',
     'Live', 'PR 12', 'Was "Out for Delivery" before PR 12 rename.'),
    ('FE-066', 'Shop Orders', 'Shop Owner',
     'Call customer',
     "Tap the customer's phone number on shop order detail to start a phone call.",
     'Shop Order detail → Customer section → tap phone number.',
     'Live', 'PR 12', 'Uses native dialer via Linking.openURL(tel:).'),
    ('FE-067', 'Shop Orders', 'Shop Owner',
     'New order alert (banner + haptic + highlight)',
     "When new orders arrive during a polling tick: banner appears at top, cards get NEW tag + primary border, phone buzzes once.",
     'Shop dashboard, automatic when new order lands.',
     'Live', 'PR 16', 'No sound yet - visual + haptic only.'),

    # ---- Shop Owner: Menu Management ----
    ('FE-070', 'Menu Management', 'Shop Owner',
     'View menu',
     'List of all menu items for your shop, organized by category.',
     'Home → Shop menu tile, or Shop dashboard → Menu.',
     'Live', '-', 'Includes both global products and custom items.'),
    ('FE-071', 'Menu Management', 'Shop Owner',
     'Add custom menu item',
     'Create a custom menu item with name, price, MRP, pack, category. Image upload currently broken.',
     'Shop menu → Add custom item button.',
     'Live', 'PR 6 (partial)', 'IMAGE UPLOAD FAILS - known issue, see FE-073.'),
    ('FE-072', 'Menu Management', 'Shop Owner',
     'Edit menu item',
     "Tap any menu item to edit price, MRP, availability, stock count.",
     'Shop menu → tap item.',
     'Live', '-', 'Server validates price + MRP relationship.'),
    ('FE-073', 'Menu Management', 'Shop Owner',
     'Image upload (KNOWN ISSUE - broken)',
     "Tapping Take Photo or Gallery results in 'storage/unauthorized' error. Use without image for now.",
     'Add/Edit menu item → Take Photo / Gallery buttons.',
     'Deferred', 'PR 6.1 parked', 'Cross-SDK auth mismatch on web/native.'),
    ('FE-074', 'Menu Management', 'Shop Owner',
     'Bulk availability toggle',
     "Select multiple items + tap Mark Available/Unavailable to flip all at once.",
     'Shop menu → tap Select → checkboxes appear → bulk action bar.',
     'Live', 'PR 8', 'Great for end-of-day restocking.'),
    ('FE-075', 'Menu Management', 'Shop Owner',
     'Stock count (optional)',
     "Per-item stock count. When stock reaches 0, item auto-marked unavailable.",
     'Edit menu item → Stock count field.',
     'Live', '-', 'Stock NOT auto-decremented on orders (offline-sales mismatch in Indian kirana).'),

    # ---- Shop Owner: Settings ----
    ('FE-080', 'Shop Settings', 'Shop Owner',
     'Edit delivery fee + minimum order',
     "Set your shop's delivery fee (in rupees) and minimum order value.",
     'Home → Shop settings tile.',
     'Live', 'PR 5', 'Validated server-side (non-negative integers).'),

    # ---- Delivery Partner ----
    ('FE-100', 'Delivery Onboarding', 'Delivery Partner',
     'Request delivery role',
     'Submit name, vehicle type (bike/scooter/cycle/walk), area. Goes to admin for approval.',
     'Home → Become Delivery Partner tile.',
     'Live', 'PR 1', 'Pending until admin approves.'),
    ('FE-101', 'Delivery Operations', 'Delivery Partner',
     'View available pickups',
     "Two sections: 'Heads up - coming soon' (accepted/preparing, can't claim yet) and 'Available now' (ready_for_pickup, claim now).",
     'Home → Delivery dashboard.',
     'Live', 'PR 12', 'Heads-up section helps partner plan routes.'),
    ('FE-102', 'Delivery Operations', 'Delivery Partner',
     'Claim a delivery',
     "Tap an Available Now order to claim. Server prevents claiming Heads-up orders (those aren't ready yet).",
     'Delivery dashboard → Available now section → tap card.',
     'Live', 'PR 12', 'Server enforces ready_for_pickup-only for claim.'),
    ('FE-103', 'Delivery Operations', 'Delivery Partner',
     'Mark picked up',
     "After claiming, tap Mark Picked Up at the shop to confirm physical handoff.",
     'My deliveries → tap claimed order → Mark Picked Up button.',
     'Live', '-', 'Sets pickedUpAt timestamp.'),
    ('FE-104', 'Delivery Operations', 'Delivery Partner',
     'Mark delivered',
     "Final action - tap when handing order to customer.",
     'My deliveries → picked-up order → Mark Delivered button.',
     'Live', '-', 'Order moves to terminal delivered state.'),

    # ---- Admin ----
    ('FE-120', 'Admin', 'Admin',
     'Admin dashboard with role tiles',
     "Home shows tiles for each admin function: All Orders, Pending Shops, Pending Delivery, User Management, Shop Management, Audit Log.",
     'Home screen as admin.',
     'Live', '-', 'Non-admin users see no admin tiles.'),
    ('FE-121', 'Admin', 'Admin',
     'View all orders across shops',
     "Live list of every order in the system, with status chip, payment status banner, delivery substates.",
     'Home → All Orders tile.',
     'Live', '-', 'Pull-to-refresh + 10s polling.'),
    ('FE-122', 'Admin', 'Admin',
     'Order full timeline disclosure',
     "Each order card has a 'Full timeline' disclosure showing every status transition with timestamp + actor.",
     'All Orders → tap any card → tap Full timeline (N steps).',
     'Live', 'PR 11', 'Reads from order.statusHistory.'),
    ('FE-123', 'Admin', 'Admin',
     'Manual override on order status',
     "Hidden by default. Admin expands disclosure to manually update an order's status - emergency lever only.",
     'All Orders → expand Manual override section on card.',
     'Live', 'PR 7', 'Writes to audit log with actorRole=admin.'),
    ('FE-124', 'Admin', 'Admin',
     'Cancel paid order + auto-refund',
     "Admin can force-cancel any paid order with a reason. Razorpay refund is auto-triggered.",
     'All Orders → order card → Cancel & Refund button.',
     'Live', 'PR 2', 'Refund usually completes in seconds.'),
    ('FE-125', 'Admin', 'Admin',
     'Pending shop approval',
     "List of shop registrations awaiting approval. View details, approve or reject (with reason).",
     'Home → Pending Shops tile.',
     'Live', '-', 'Approve grants shopOwner claim with shopId.'),
    ('FE-126', 'Admin', 'Admin',
     'Pending delivery partner approval',
     "List of delivery role requests. View applicant details, approve or reject.",
     'Home → Pending Delivery Requests tile.',
     'Live', 'PR 1', 'Approve grants delivery claim.'),
    ('FE-127', 'Admin', 'Admin',
     'User management with role revocation',
     "View all users + their roles. Revoke shopOwner / delivery / admin claims.",
     'Home → User Management tile.',
     'Live', '-', 'Revocation logs to audit.'),
    ('FE-128', 'Admin', 'Admin',
     'Shop management (suspend/unsuspend)',
     "View all active shops. Suspend a shop (hides from customers) or unsuspend (restores).",
     'Home → Shop Management tile.',
     'Live', '-', 'Suspended shops invisible to non-admin/non-owner.'),
    ('FE-129', 'Admin', 'Admin',
     'Edit shop settings on behalf of shop',
     "Admin can edit any shop's delivery fee + minimum order without being the shop owner.",
     'Shop Management → tap shop → Edit settings.',
     'Live', 'PR 5', 'Useful when shop owners are unreachable.'),
    ('FE-130', 'Admin', 'Admin',
     'Audit log viewer',
     "Chronological log of admin actions (approvals, revocations, suspensions, refunds, cancellations).",
     'Home → Audit Log tile.',
     'Live', 'PR 8', 'Customer self-cancellations now logged as actorRole=customer (PR 8.1).'),

    # ---- Cross-cutting / System ----
    ('FE-200', 'System', 'All Users',
     'Sign out clears local state',
     'Signing out wipes cart, persisted auth, and any in-memory stores.',
     'Profile → Sign Out.',
     'Live', '-', 'Per Phase 3 cleanup contract.'),
    ('FE-201', 'System', 'All Users',
     'Sentry error reporting',
     "Crashes and uncaught errors are reported to Sentry for monitoring.",
     'Automatic on any unhandled exception.',
     'Live', '-', 'Sentry DSN configured in app.json.'),
    ('FE-202', 'System', 'All Users',
     'Order auto-cancel on payment timeout',
     "Online orders where payment didn't complete within 24h are auto-cancelled by a daily cron.",
     'Automatic. Customer sees cancelled status next time they look.',
     'Live', 'PR 3', 'cleanupAbandonedOrders runs once daily.'),
    ('FE-203', 'System', 'All Users',
     'Razorpay webhook for payment confirmation',
     'Razorpay calls our webhook to confirm captured payments. Order auto-transitions to paid.',
     'Automatic on payment success.',
     'Live', 'PR 2', 'HMAC-signed for security.'),
    ('FE-204', 'System', 'All Users',
     'Cross-platform: web preview',
     "App runs in browsers via expo export --platform web for quick UX preview. Native is the production target.",
     'Currently paused - web hosting setup not pursued.',
     'Deferred', '-', 'Native (iOS/Android) is the primary delivery surface.'),
]


# ============================================================================
# TEST CASES
# (TC ID, Feature ID, Role, Title, Type, Priority,
#  Pre-conditions, Steps, Test Data, Expected Result)
# ============================================================================
TEST_CASES = [
    # ---- Authentication ----
    ('TC-001', 'FE-001', 'All', 'Login with valid test phone number',
     'Positive', 'Critical',
     'App is signed out. Test phone +91 99999 99991 is configured in Firebase Console with OTP 123456.',
     '1. Open app\n2. On Sign in screen, enter 9999999991\n3. Tap Send OTP\n4. On OTP screen, enter 123456\n5. Tap Verify',
     'Phone: 9999999991, OTP: 123456',
     'User is signed in. Lands on Home screen (or completes profile setup if first time).'),
    ('TC-002', 'FE-001', 'All', 'Login with invalid OTP shows error',
     'Negative', 'High',
     'Signed out. Test phone configured.',
     '1. Sign in flow, enter 9999999991\n2. Tap Send OTP\n3. Enter 000000 (wrong)\n4. Tap Verify',
     'OTP: 000000',
     'Error message: "Invalid OTP. Try again." User stays on OTP screen.'),
    ('TC-003', 'FE-001', 'All', 'Phone with less than 10 digits disables Send OTP',
     'Negative', 'Medium',
     'On Sign in screen.',
     '1. Type 99999 (5 digits)\n2. Observe Send OTP button',
     'Phone: 99999',
     'Send OTP button is disabled (greyed out) until 10 digits entered.'),
    ('TC-004', 'FE-002', 'All', 'Resend OTP shows countdown after Send',
     'Positive', 'High',
     'Just tapped Send OTP and arrived at OTP screen.',
     '1. Look at Resend OTP link\n2. Wait 5 seconds',
     '-',
     "Initially shows 'Resend OTP in 30s' (greyed out). Countdown ticks down each second."),
    ('TC-005', 'FE-002', 'All', 'Resend OTP after cooldown sends new OTP',
     'Positive', 'High',
     'On OTP screen, cooldown has reached 0.',
     "1. Wait for countdown to reach 0\n2. Link becomes 'Didn\\'t get the code? Resend OTP'\n3. Tap it",
     '-',
     'New OTP request fires. Cooldown restarts at 30s. Test phone receives same code (123456); real phones get a fresh SMS.'),
    ('TC-006', 'FE-002', 'All', 'Resend OTP rate-limit error message',
     'Negative', 'Medium',
     'Trigger by tapping Resend many times in succession on a real phone.',
     '1. Resend OTP 3-4 times in 2 minutes\n2. Observe error',
     '-',
     "Error: 'Too many OTP requests for this number. Wait a few minutes and try again.'"),
    ('TC-007', 'FE-003', 'All', 'New user must enter full name',
     'Negative', 'High',
     'Just verified OTP for a brand new phone number.',
     '1. After OTP verification, lands on Profile setup\n2. Leave Name field empty\n3. Tap Save',
     'Name: (empty)',
     'Validation error: "Name required". Save button stays disabled or alert appears.'),
    ('TC-008', 'FE-003', 'All', 'Name with only spaces is rejected',
     'Edge case', 'Medium',
     'Brand new user on Profile setup.',
     '1. Enter "   " (only spaces) in Name field\n2. Tap Save',
     'Name: "   "',
     'Validation rejects whitespace-only name.'),
    ('TC-009', 'FE-003', 'All', 'Phone number field is read-only',
     'Edge case', 'Medium',
     'On Profile setup screen.',
     '1. Try to edit phone number field',
     '-',
     'Phone field is non-editable (read-only display). Pre-filled with verified phone.'),
    ('TC-010', 'FE-004', 'All', 'Sign out clears cart',
     'Positive', 'Critical',
     'Signed in with items in cart.',
     '1. Verify cart has items\n2. Go to Profile\n3. Tap Sign Out\n4. Sign back in (same or different account)',
     '-',
     'After sign-out + sign-in, cart is empty. Local persisted cart state was wiped.'),
    ('TC-011', 'FE-005', 'All', 'Multi-role user sees Your Roles section',
     'Positive', 'High',
     'Signed in as a user who has shopOwner + delivery + admin claims.',
     '1. Sign in\n2. Look at Home screen',
     '-',
     "'Your Roles' section visible with Shop Dashboard, Delivery Dashboard, Admin tiles."),
    ('TC-012', 'FE-006', 'Customer', 'Save a new delivery address',
     'Positive', 'High',
     'Signed in customer with no saved addresses.',
     '1. Profile → Saved Addresses → Add\n2. Fill Name, House/Flat, Street, City, Pincode, Phone\n3. Save',
     'Name: Home, City: Bangalore, Pincode: 560001',
     "Address appears in Saved Addresses list. Used as default if it's the first."),

    # ---- Browse & Search ----
    ('TC-020', 'FE-010', 'Customer', 'Browse shops shows all active shops',
     'Positive', 'High',
     'Logged in customer. Multiple shops exist in dev project.',
     '1. Home → Browse shops near me\n2. Wait for list to load',
     '-',
     "All active shops appear (no distance filter in testing mode). Each card shows shop name + 'Order again' or browse CTA."),
    ('TC-021', 'FE-010', 'Customer', 'Suspended shop is hidden',
     'Negative', 'High',
     'Admin has suspended Shop X. Sign in as customer.',
     '1. Browse shops list',
     '-',
     'Shop X not visible.'),
    ('TC-022', 'FE-011', 'Customer', 'Open shop detail and see menu',
     'Positive', 'Critical',
     'Active shop with menu items.',
     '1. Browse → tap a shop card\n2. Scroll through menu',
     '-',
     'Menu loads with items grouped by category. Each item shows name, price, MRP, pack label, +/- buttons.'),
    ('TC-023', 'FE-012', 'Customer', 'Search "atta" returns relevant items',
     'Positive', 'High',
     'Shops have atta in menu.',
     '1. Home → tap Search bar\n2. Type atta\n3. View results',
     'Query: atta',
     'Results list atta variants from any shop. Each result shows item name + shop name.'),
    ('TC-024', 'FE-012', 'Customer', 'Search with no results shows empty state',
     'Negative', 'Medium',
     'On Search.',
     '1. Search for "xyznonexistent"',
     '-',
     'Empty state message: "No items found for ___". No crash.'),
    ('TC-025', 'FE-013', 'Customer', 'Filter search by category',
     'Edge case', 'Medium',
     'On Search.',
     '1. Tap a category chip (e.g. Beverages)\n2. Observe filter applied',
     '-',
     'Results filtered to that category. Selected chip is visually active.'),

    # ---- Cart ----
    ('TC-030', 'FE-020', 'Customer', 'Add item to cart from shop detail',
     'Positive', 'Critical',
     'On a shop detail screen with menu items.',
     '1. Tap + on any item\n2. Observe cart icon',
     '-',
     'Item quantity becomes 1. Bottom cart bar shows "View cart - rupees X". Cart icon badge increments.'),
    ('TC-031', 'FE-021', 'Customer', 'Decrement to zero removes item from cart',
     'Edge case', 'High',
     'Cart has one item with quantity 1.',
     '1. Open Cart screen\n2. Tap - on that item',
     '-',
     'Item removed from cart. If cart now empty, cart shows empty state and shop binding is cleared.'),
    ('TC-032', 'FE-022', 'Customer', 'Adding from different shop shows Replace cart prompt',
     'Negative', 'Critical',
     'Cart has items from Shop A.',
     '1. Browse to Shop B\n2. Tap + on any Shop B item',
     '-',
     "Alert: 'Your cart has items from Shop A. Replace with Shop B items?' Cancel = no change. Replace = cart cleared then Shop B item added."),
    ('TC-033', 'FE-023', 'Customer', 'Cart survives app force-close',
     'Positive', 'High',
     'Cart has 3 items.',
     '1. Note cart contents\n2. Force-close app\n3. Reopen app and view cart',
     '-',
     'All 3 items still in cart with original quantities. Total recomputes correctly.'),

    # ---- Checkout & Payment ----
    ('TC-040', 'FE-030', 'Customer', 'Address picker shows saved addresses',
     'Positive', 'High',
     'Customer with 2 saved addresses, cart ready.',
     '1. Cart → Proceed to checkout\n2. Tap address picker',
     '-',
     'Both saved addresses listed. Tapping one selects it. Option to Add new address present.'),
    ('TC-041', 'FE-031', 'Customer', 'Choose COD as payment method',
     'Positive', 'Critical',
     'On checkout screen with address selected.',
     '1. Tap Cash on Delivery radio\n2. Tap Place Order',
     '-',
     'Order is created. Navigates to Order Confirmation / Order Detail screen with status Pending.'),
    ('TC-042', 'FE-032', 'Customer', 'Pay online with Razorpay test card (success)',
     'Positive', 'Critical',
     'Razorpay test keys configured. On checkout.',
     '1. Tap Online (Razorpay)\n2. Tap Place Order\n3. In Razorpay overlay, choose Card\n4. Enter test card 4386 2894 0766 0153\n5. CVV: 123, Expiry: 12/30\n6. Submit, complete 3DS with OTP 1234',
     'Card: 4386 2894 0766 0153, OTP: 1234',
     'Razorpay confirms. Order transitions paymentStatus to paid within ~10s (webhook). Status visible on Order Detail.'),
    ('TC-043', 'FE-032', 'Customer', 'International card is rejected by Razorpay',
     'Negative', 'High',
     'On Razorpay overlay.',
     '1. Enter 4111 1111 1111 1111 (global Visa test)\n2. Submit',
     'Card: 4111 1111 1111 1111',
     'Razorpay rejects with "International cards not accepted". Use Indian test cards only.'),
    ('TC-044', 'FE-032', 'Customer', 'Pay with test UPI success@razorpay',
     'Positive', 'High',
     'UPI enabled in Razorpay test dashboard.',
     '1. On Razorpay overlay, tap UPI\n2. Enter success@razorpay\n3. Submit',
     'UPI: success@razorpay',
     'Instant success. Order to paid.'),
    ('TC-045', 'FE-033', 'Customer', 'Dismiss Razorpay leaves order in pending payment',
     'Negative', 'High',
     'On checkout for online payment.',
     '1. Tap Place Order to open Razorpay\n2. Tap X / dismiss overlay without paying',
     '-',
     'Order created with paymentStatus pending. Order detail shows "Payment incomplete" with Pay Now / Cancel buttons.'),
    ('TC-046', 'FE-033', 'Customer', 'Retry payment on pending order',
     'Positive', 'High',
     'Order in pending payment state.',
     '1. Order detail → Tap Pay Now\n2. Complete Razorpay flow',
     '-',
     'Order transitions to paid.'),
    ('TC-047', 'FE-034', 'Customer', 'Minimum order warning below threshold',
     'Negative', 'High',
     'Shop has minimum order 200. Cart subtotal 150.',
     '1. Open Cart\n2. Observe checkout area',
     '-',
     'Warning: "Add 50 more to proceed" (or similar). Proceed to checkout button disabled or warns.'),

    # ---- Customer Orders ----
    ('TC-050', 'FE-040', 'Customer', 'Active orders rail appears on Home after placing order',
     'Positive', 'Critical',
     'Customer with no in-flight orders. Places an order.',
     '1. Place an order\n2. Navigate back to Home\n3. Look at top of screen',
     '-',
     "'Your active orders' rail visible with one card showing shop name, Pending chip, ETA."),
    ('TC-051', 'FE-040', 'Customer', 'Active orders rail auto-hides when no in-flight orders',
     'Edge case', 'High',
     'New customer with no orders.',
     '1. Sign in fresh\n2. Look at Home',
     '-',
     'Active orders rail not visible. Only search + categories + (eventually) Order Again rail.'),
    ('TC-052', 'FE-040', 'Customer', 'Multiple active orders sort newest-first',
     'Edge case', 'Medium',
     'Customer with 3 active orders from 3 shops.',
     '1. Place 3 orders in sequence\n2. View Home',
     '-',
     'All 3 cards in rail. Leftmost is the most recently placed.'),
    ('TC-053', 'FE-040', 'Customer', 'Tap active order card opens detail',
     'Positive', 'High',
     'Active orders rail visible.',
     '1. Tap an active card',
     '-',
     'Navigates to OrderDetail with full info.'),
    ('TC-054', 'FE-041', 'Customer', 'Order Again rail appears after first delivered order',
     'Positive', 'High',
     'Customer with 1 delivered order, no active.',
     '1. Sign in\n2. Home',
     '-',
     "'Order again' rail visible with one card for that shop."),
    ('TC-055', 'FE-041', 'Customer', 'Order Again rail ranks by frequency',
     'Edge case', 'Medium',
     'Customer with 3 delivered orders from Shop A, 1 from Shop B.',
     '1. Sign in\n2. Home',
     '-',
     'Shop A appears first (more orders), Shop B second.'),
    ('TC-056', 'FE-041', 'Customer', "Cancelled orders don't appear in Order Again",
     'Edge case', 'Medium',
     'Customer cancelled an order from Shop C, never had a delivered one from Shop C.',
     '1. Sign in\n2. Home',
     '-',
     'Shop C not in Order Again rail.'),
    ('TC-057', 'FE-042', 'Customer', 'Reorder from a delivered order with all items available',
     'Positive', 'Critical',
     'Customer with delivered order from Shop A. All items still in shop menu at same prices.',
     '1. My Orders → tap Reorder on that card\n2. Modal opens with item list\n3. Tap "Add N items to cart"',
     '-',
     'Modal shows all items as available, same prices. Cart fills with original quantities. Navigates to Cart.'),
    ('TC-058', 'FE-042', 'Customer', 'Reorder with price change shows old vs new',
     'Edge case', 'High',
     'Past order had atta at 250. Shop changed it to 275.',
     '1. Reorder modal opens for that order',
     '-',
     'Atta line shows current price 275 with old 250 struck through and a "+10%" badge.'),
    ('TC-059', 'FE-042', 'Customer', 'Reorder with some items unavailable',
     'Edge case', 'High',
     'Past order had 5 items. Shop marked 2 as unavailable.',
     '1. Reorder modal opens',
     '-',
     'Modal shows 3 items in "Available" section, 2 in "Unavailable" with reason. CTA: "Add 3 items to cart".'),
    ('TC-060', 'FE-042', 'Customer', 'Reorder when all items unavailable',
     'Edge case', 'Medium',
     'Shop suspended every item from past order.',
     '1. Reorder modal opens',
     '-',
     'All items in Unavailable. CTA disabled with text "No items available".'),
    ('TC-061', 'FE-042', 'Customer', 'Reorder when shop is suspended',
     'Negative', 'High',
     'Past order from Shop X. Admin suspended Shop X.',
     '1. My Orders → tap Reorder on Shop X order',
     '-',
     'Modal closes. Alert: "This shop is no longer accepting orders. Try a different shop." Cart unchanged.'),
    ('TC-062', 'FE-042', 'Customer', 'Reorder replaces existing cart from different shop',
     'Edge case', 'High',
     'Cart has items from Shop A. Customer reorders from Shop B.',
     '1. Cart has Shop A items\n2. My Orders → Reorder on a Shop B past order\n3. Confirm modal',
     '-',
     'Cart cleared of Shop A items, refilled with Shop B items. Navigates to Cart.'),
    ('TC-063', 'FE-042', 'Customer', 'Reorder works for cancelled orders too',
     'Edge case', 'Medium',
     'Customer cancelled a paid order from Shop X (within 2-min window).',
     '1. My Orders → find cancelled order → tap Reorder',
     '-',
     'Modal opens normally. Reorder works identically.'),
    ('TC-064', 'FE-043', 'Customer', 'Order detail shows all fields',
     'Positive', 'Critical',
     'Customer with at least one order.',
     '1. My Orders → tap any order',
     '-',
     'Visible: order ID, placed time, status chip, items list, delivery address, payment method, payment status, action buttons (where applicable).'),
    ('TC-065', 'FE-044', 'Customer', 'Cancel pending payment-failed order',
     'Positive', 'High',
     'Order with paymentStatus pending (Razorpay was dismissed).',
     '1. Order detail → Tap Cancel order',
     '-',
     'Order status to cancelled. No charge to customer.'),
    ('TC-066', 'FE-045', 'Customer', 'Cancel paid order within 2-min window',
     'Positive', 'Critical',
     'Just placed and paid for an online order (< 2 min ago).',
     '1. Order detail shows "Cancel order (X:XX left)" green button\n2. Tap it\n3. Confirm in alert',
     '-',
     'Order to cancelled. Payment Status to Refund pending, then to Refunded within ~10s.'),
    ('TC-067', 'FE-045', 'Customer', 'Cancel button hidden after 2 min',
     'Negative', 'High',
     'Paid order placed > 2 min ago.',
     '1. Open order detail\n2. Look for Cancel option',
     '-',
     "Green Cancel button not visible. Instead shows 'Cancellation window expired. Contact support if you still need to cancel.'"),
    ('TC-068', 'FE-045', 'Customer', 'Cancel countdown ticks live',
     'Edge case', 'Medium',
     'Order placed < 2 min ago. Stay on detail screen.',
     '1. Open order detail\n2. Observe countdown',
     '-',
     'Cancel button label shows e.g. "Cancel order (1:32 left)" decrementing each second.'),
    ('TC-069', 'FE-046', 'Customer', 'Refund status displays after cancellation',
     'Positive', 'High',
     'Cancelled a paid order via 2-min window 30s ago.',
     '1. Order detail → Payment section',
     '-',
     'Shows "Refunded" in green (refund usually completes in seconds). Note text: "Funds typically reach your account in 5-7 business days."'),
    ('TC-070', 'FE-047', 'Customer', 'Customer sees "Out for delivery" not "Ready for Pickup"',
     'Edge case', 'High',
     'Shop has marked order Ready for Pickup. Customer views order.',
     '1. Customer → Order detail or Active rail on Home',
     '-',
     'Status chip shows "Out for delivery" (customer-facing label). Internally the status is ready_for_pickup.'),

    # ---- Shop Owner: Registration ----
    ('TC-080', 'FE-050', 'Shop Owner', 'Register a new shop',
     'Positive', 'Critical',
     'User has no shop. On Home.',
     '1. Home → Become a Shop Owner\n2. Fill all required fields\n3. Submit',
     'Shop name: Test Kirana, City: Bangalore',
     'Shop submitted with status pending. Navigates to Waiting for Approval screen.'),
    ('TC-081', 'FE-051', 'Shop Owner', 'Pending shop tile shows on Home',
     'Positive', 'High',
     'Just submitted shop registration.',
     '1. Navigate back to Home\n2. Look for tile',
     '-',
     "'Awaiting approval' tile visible. Tap to re-open Waiting for Approval screen with current status."),

    # ---- Shop Owner: Dashboard & Orders ----
    ('TC-090', 'FE-060', 'Shop Owner', 'Shop dashboard loads with active orders',
     'Positive', 'Critical',
     'Approved shop with orders.',
     '1. Home → Shop dashboard tile',
     '-',
     'List of active orders visible. Each card shows order ID, customer name, items count, total, status chip.'),
    ('TC-091', 'FE-061', 'Shop Owner', 'Pull-to-refresh updates the list',
     'Positive', 'High',
     'On shop dashboard.',
     '1. Pull down on the list\n2. Release',
     '-',
     'Loader appears briefly. List refreshes with latest orders.'),
    ('TC-092', 'FE-062', 'Shop Owner', "Today's stats card shows correct counts",
     'Positive', 'High',
     'Shop has 5 orders today: 3 active, 1 delivered, 1 cancelled.',
     '1. Look at stats card on top of dashboard',
     '-',
     "Pending count = 3 (active), today's total = 5, revenue = sum of paid orders today."),
    ('TC-093', 'FE-063', 'Shop Owner', 'Accept order with valid ETA',
     'Positive', 'Critical',
     'Pending order in dashboard.',
     '1. Tap order card → Order detail\n2. Tap Accept button\n3. ETA modal opens, enter 20\n4. Tap Accept (inside modal)',
     'ETA: 20 min',
     'Order status to accepted. ETA field populated. Customer sees "Arriving in ~20 min". Delivery partners see in Heads-up.'),
    ('TC-094', 'FE-063', 'Shop Owner', 'Accept without ETA blocked server-side',
     'Negative', 'High',
     'Shop tries to call updateOrderStatus directly without readyByEstimate.',
     '1. (Backend) Call accept without ETA',
     '-',
     'Server rejects with invalid-argument: "ETA required when accepting an order".'),
    ('TC-095', 'FE-063', 'Shop Owner', 'Accept with ETA = 0 rejected',
     'Negative', 'Medium',
     'On Accept modal.',
     '1. Enter 0 in minutes field\n2. Tap Accept',
     'ETA: 0',
     'Alert: "Invalid ETA. Enter a number of minutes between 1 and 240."'),
    ('TC-096', 'FE-063', 'Shop Owner', 'Accept with ETA = 240 (max) works',
     'Edge case', 'Low',
     'On Accept modal.',
     '1. Enter 240\n2. Submit',
     'ETA: 240 (4 hours)',
     'Order accepted with 4-hour ETA. Unusual but valid.'),
    ('TC-097', 'FE-064', 'Shop Owner', 'Update ETA on Start Preparing',
     'Positive', 'High',
     'Accepted order with 20 min ETA.',
     '1. Order detail → Start Preparing\n2. Modal opens prefilled with existing ETA\n3. Change to 35\n4. Submit',
     'New ETA: 35 min',
     'Order to preparing. ETA updated. Admin dashboard shows "updated from".'),
    ('TC-098', 'FE-065', 'Shop Owner', 'Mark Ready for Pickup',
     'Positive', 'Critical',
     'Order in preparing state.',
     '1. Order detail → Ready for Pickup button\n2. Confirm',
     '-',
     'Order to ready_for_pickup. Appears in Available now section of delivery dashboard.'),
    ('TC-099', 'FE-066', 'Shop Owner', 'Call customer from order detail',
     'Positive', 'Medium',
     'On shop order detail.',
     '1. Tap customer phone number',
     '-',
     'Native phone dialer opens with number pre-filled.'),
    ('TC-100', 'FE-067', 'Shop Owner', 'New order alert banner appears',
     'Positive', 'Critical',
     'Shop owner on dashboard. Have a customer-tester place an order.',
     '1. Stay on shop dashboard\n2. Customer-tester places order\n3. Wait up to 10s (polling cycle)',
     '-',
     "Banner appears at top: '1 new order'. Phone buzzes once. Order card has green border + 'NEW' tag."),
    ('TC-101', 'FE-067', 'Shop Owner', 'First dashboard open shows no false-positive new alerts',
     'Edge case', 'High',
     'Fresh app launch with pre-existing orders.',
     '1. Sign in fresh\n2. Open Shop dashboard',
     '-',
     'No banner, no haptic, no NEW tags. First tick silently establishes baseline.'),
    ('TC-102', 'FE-067', 'Shop Owner', 'Tap order card clears NEW highlight',
     'Positive', 'High',
     'Banner showing, one new card.',
     '1. Tap the NEW-tagged card\n2. Navigate back',
     '-',
     'Banner gone. No NEW tags remaining.'),
    ('TC-103', 'FE-067', 'Shop Owner', 'Multiple new orders show single haptic',
     'Edge case', 'Medium',
     'Three orders placed simultaneously by 3 testers.',
     '1. Watch dashboard\n2. Wait for next polling tick',
     '-',
     "Banner: '3 new orders'. All 3 cards have NEW tags. Phone buzzes ONCE, not three times."),

    # ---- Shop Owner: Menu Management ----
    ('TC-110', 'FE-070', 'Shop Owner', 'View menu list',
     'Positive', 'High',
     'Shop with at least 5 menu items.',
     '1. Home → Shop menu\n2. Scroll',
     '-',
     'Items load grouped by category. Each shows name, price, MRP, availability badge.'),
    ('TC-111', 'FE-071', 'Shop Owner', 'Add a custom menu item (without image)',
     'Positive', 'High',
     'On shop menu.',
     '1. Add custom item button\n2. Enter Name: Fresh paneer 250g, Price: 80, MRP: 100, Pack: 250 g, Category: pick one\n3. Skip image (broken)\n4. Save',
     'Name: Fresh paneer 250g, Price: 80, MRP: 100',
     'Item added to menu. Visible in list.'),
    ('TC-112', 'FE-072', 'Shop Owner', 'Edit a menu item price',
     'Positive', 'High',
     'Existing menu item.',
     '1. Tap item → Edit\n2. Change price\n3. Save',
     'New price: 95',
     'Item updated. Customers will see new price next browse.'),
    ('TC-113', 'FE-073', 'Shop Owner', 'Image upload fails (KNOWN ISSUE)',
     'Negative', 'Medium',
     'On Add/Edit menu item.',
     '1. Tap Take Photo or Gallery\n2. Pick an image\n3. Observe upload',
     '-',
     "Upload fails with 'storage/unauthorized' error. Save item without image for now."),
    ('TC-114', 'FE-074', 'Shop Owner', 'Bulk mark 5 items unavailable',
     'Positive', 'High',
     'Shop menu with 10+ items.',
     '1. Shop menu → tap Select (multi-select mode)\n2. Check 5 items\n3. Tap Mark Unavailable in bottom bar',
     '-',
     'All 5 items flip to unavailable. Customers can no longer add them.'),
    ('TC-115', 'FE-074', 'Shop Owner', 'Bulk action with empty selection no-op',
     'Edge case', 'Low',
     'Multi-select mode, nothing checked.',
     '1. Tap Mark Available without selecting anything',
     '-',
     'Button disabled, or no-op. No items modified.'),

    # ---- Shop Settings ----
    ('TC-120', 'FE-080', 'Shop Owner', 'Update delivery fee + minimum order',
     'Positive', 'High',
     'Shop owner on Shop Settings.',
     '1. Home → Shop Settings\n2. Change delivery fee to 30\n3. Change min order to 200\n4. Save',
     'Delivery fee: 30, Min order: 200',
     'Settings saved. Customer checkout uses new values. Audit log entry created.'),
    ('TC-121', 'FE-080', 'Shop Owner', 'Negative delivery fee rejected',
     'Negative', 'Medium',
     'On Shop Settings.',
     '1. Enter -10 for delivery fee\n2. Save',
     'Delivery fee: -10',
     'Validation error. Save blocked.'),

    # ---- Delivery Partner ----
    ('TC-130', 'FE-100', 'Delivery Partner', 'Request delivery role',
     'Positive', 'High',
     'User with no delivery role.',
     '1. Home → Become Delivery Partner\n2. Fill name, vehicle, area\n3. Submit',
     'Vehicle: Bike, Area: Indiranagar',
     'Request submitted with status pending. Goes to admin queue.'),
    ('TC-131', 'FE-101', 'Delivery Partner', 'See Heads up + Available now sections',
     'Positive', 'Critical',
     'Approved delivery partner. Shop has 1 accepted order + 1 ready_for_pickup order in same area.',
     '1. Home → Delivery dashboard\n2. Observe sections',
     '-',
     "'Heads up - coming soon' shows accepted order with 'Ready by X PM' badge. 'Available now' shows ready_for_pickup. Both visible."),
    ('TC-132', 'FE-101', 'Delivery Partner', 'Heads up card is informational only',
     'Edge case', 'Medium',
     'On delivery dashboard.',
     '1. Tap a Heads up card',
     '-',
     'Card opens detail (read-only). No Claim button visible OR Claim is disabled with explanation.'),
    ('TC-133', 'FE-102', 'Delivery Partner', 'Claim an Available Now order',
     'Positive', 'Critical',
     'Delivery dashboard with 1 ready_for_pickup order.',
     '1. Tap an Available now card\n2. Tap Claim',
     '-',
     'Order claimed, deliveryPersonId set to this partner. Moves to My Deliveries.'),
    ('TC-134', 'FE-103', 'Delivery Partner', 'Mark Picked Up',
     'Positive', 'High',
     'Claimed order, partner at shop.',
     '1. My Deliveries → tap claimed order\n2. Mark Picked Up',
     '-',
     'pickedUpAt timestamp set. Customer card on Home shows "Out for delivery".'),
    ('TC-135', 'FE-104', 'Delivery Partner', 'Mark Delivered',
     'Positive', 'Critical',
     'Picked-up order, partner at customer location.',
     '1. Tap Mark Delivered\n2. Confirm',
     '-',
     'Order to delivered. Customer card disappears from active rail, shop appears in Order Again rail.'),

    # ---- Admin ----
    ('TC-150', 'FE-120', 'Admin', 'Admin home shows admin tiles',
     'Positive', 'High',
     'Signed in as admin.',
     '1. Home',
     '-',
     'Visible: All Orders, Pending Shops, Pending Delivery, User Management, Shop Management, Audit Log tiles.'),
    ('TC-151', 'FE-121', 'Admin', 'All Orders shows orders from all shops',
     'Positive', 'Critical',
     'Multiple shops with orders.',
     '1. Home → All Orders',
     '-',
     'List of every order. Each card shows order ID, customer, shop name, status chip, total.'),
    ('TC-152', 'FE-122', 'Admin', 'Expand order timeline disclosure',
     'Positive', 'High',
     'On All Orders, pick an order with 4+ status transitions.',
     '1. Tap "Full timeline (N steps)" on a card',
     '-',
     'Timeline expands showing each transition with timestamp, actor (role + truncated uid), and reason if present.'),
    ('TC-153', 'FE-122', 'Admin', 'Customer-cancelled order shows actorRole=customer',
     'Edge case', 'Medium',
     'Customer cancelled a paid order via 2-min window.',
     '1. Admin → All Orders → find that order → expand timeline',
     '-',
     'Cancellation entry shows "by customer:XXXX..." not "by system:...".'),
    ('TC-154', 'FE-123', 'Admin', 'Manual override is hidden by default',
     'Positive', 'Medium',
     'On any order card.',
     '1. Look at card',
     '-',
     'No action buttons visible by default. "Manual override" disclosure present.'),
    ('TC-155', 'FE-123', 'Admin', 'Expanding manual override reveals status buttons',
     'Positive', 'High',
     'On any order card.',
     '1. Tap "Manual override"',
     '-',
     "Status transition buttons appear (Accept, Mark Preparing, etc.). Warning text reminds it's an emergency lever."),
    ('TC-156', 'FE-124', 'Admin', 'Cancel + refund a paid order',
     'Positive', 'Critical',
     'Paid online order in any state.',
     '1. All Orders → tap order → Cancel & Refund\n2. Enter reason\n3. Submit',
     'Reason: Customer requested',
     'Order to cancelled. paymentStatus to refunded within ~10s. Audit log entry.'),
    ('TC-157', 'FE-125', 'Admin', 'Approve a pending shop',
     'Positive', 'Critical',
     'Pending shop in queue.',
     '1. Home → Pending Shops → tap shop\n2. Review details\n3. Tap Approve',
     '-',
     'Shop status to active. Owner gets shopOwner + shopId claims. Visible to customers.'),
    ('TC-158', 'FE-125', 'Admin', 'Reject a pending shop with reason',
     'Positive', 'High',
     'Pending shop.',
     '1. Pending Shops → tap shop → Reject\n2. Enter reason\n3. Confirm',
     'Reason: Address unclear',
     'Shop status to rejected. Owner sees rejection reason on Waiting for Approval screen.'),
    ('TC-159', 'FE-126', 'Admin', 'Approve a delivery role request',
     'Positive', 'High',
     'Pending delivery request in queue.',
     '1. Home → Pending Delivery → tap request\n2. Approve',
     '-',
     'User gets delivery claim. Delivery dashboard becomes available to them.'),
    ('TC-160', 'FE-127', 'Admin', 'Revoke shopOwner role',
     'Positive', 'High',
     'Admin in User Management. Target user has shopOwner claim.',
     '1. User Management → tap target user → Revoke shopOwner',
     '-',
     "shopOwner claim removed. User's shop dashboard becomes inaccessible after token refresh. Audit log entry."),
    ('TC-161', 'FE-128', 'Admin', 'Suspend a shop',
     'Positive', 'High',
     'Active shop visible to customers.',
     '1. Shop Management → tap shop → Suspend\n2. Confirm',
     '-',
     'Shop status to suspended. Hidden from customer browse and search. Admin can unsuspend later.'),
    ('TC-162', 'FE-129', 'Admin', 'Edit settings for any shop',
     'Positive', 'High',
     "Admin on a shop's detail page in Shop Management.",
     '1. Tap "Edit settings"\n2. Change delivery fee\n3. Save',
     '-',
     'Settings saved on that shop. Audit log entry with actorRole=admin.'),
    ('TC-163', 'FE-130', 'Admin', 'Audit log shows recent admin actions',
     'Positive', 'Critical',
     'Multiple admin actions performed in last hour.',
     '1. Home → Audit Log\n2. Scroll',
     '-',
     'Chronological list. Each entry shows timestamp, actor (role + uid), action type, target.'),
    ('TC-164', 'FE-130', 'Admin', 'Non-admin cannot access audit log',
     'Negative', 'High',
     'Sign in as customer (no admin claim).',
     '1. Try to call listRecentAuditEntries via app\n(no UI path; verify server rejects)',
     '-',
     'Server returns permission-denied. UI never exposes the tile to non-admins.'),

    # ---- Cross-cutting & smoke ----
    ('TC-180', 'FE-200', 'All', 'Sign out + sign back in clears cart',
     'Positive', 'High',
     'Customer with items in cart.',
     '1. Sign out\n2. Sign in as same or different user',
     '-',
     'Cart is empty. Old cart was wiped.'),
    ('TC-181', 'FE-202', 'All', 'Abandoned online order auto-cancelled after 24h',
     'Positive', 'Medium',
     'Customer placed online order, dismissed Razorpay. Order in pending payment for >24h.',
     '1. (Wait 24h for cron, or trigger manually via admin)\n2. Check order status',
     '-',
     'Order to cancelled, paymentStatus to expired. Audit log entry actorRole=system.'),
    ('TC-182', '-', 'All', 'App handles network loss gracefully',
     'Edge case', 'High',
     'On any screen with data fetch.',
     '1. Enable airplane mode\n2. Navigate around the app',
     '-',
     'Screens show error banners or stale data. No crashes. Retry buttons work when network returns.'),
    ('TC-183', '-', 'All', 'No screen crashes (ErrorBoundary check)',
     'Negative', 'Critical',
     "Smoke test all major screens.",
     '1. Visit every reachable screen as each role\n2. Check Sentry afterward',
     '-',
     'No ErrorBoundary "Something went wrong" screens encountered. Sentry shows no JS exceptions.'),
    ('TC-184', '-', 'All', 'Force-close + reopen preserves session',
     'Edge case', 'High',
     'Signed in.',
     '1. Force-close app\n2. Reopen',
     '-',
     'Still signed in. Cart preserved. Lands on Home.'),
    ('TC-185', '-', 'Customer', 'Place order end-to-end (COD smoke)',
     'Positive', 'Critical',
     'Customer with saved address.',
     '1. Browse → pick shop → add items meeting min order\n2. Cart → Checkout\n3. Select COD, Place Order\n4. View Order Detail\n5. As Shop, accept with ETA\n6. As Shop, mark preparing, then ready_for_pickup\n7. As Delivery, claim, pick up, deliver\n8. As Customer, return to app',
     '-',
     'Order completes full lifecycle. Customer Active rail empties; Order Again rail gains that shop.'),
    ('TC-186', '-', 'Customer', 'Place order end-to-end (Online smoke)',
     'Positive', 'Critical',
     'Customer with saved address. Razorpay test keys.',
     '1. Same as TC-185 but choose Online payment\n2. Pay with 4386 2894 0766 0153, OTP 1234\n3. Wait for paid status\n4. Continue lifecycle',
     '-',
     'Payment captured via webhook. paymentStatus to paid. Rest of lifecycle as TC-185.'),
]


# ============================================================================
# WORKBOOK BUILDERS
# ============================================================================

def style_header_row(sheet, num_cols):
    for col_idx in range(1, num_cols + 1):
        cell = sheet.cell(row=1, column=col_idx)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')
        cell.border = BORDER
    sheet.row_dimensions[1].height = 32
    sheet.freeze_panes = 'A2'


def style_body_cells(sheet, num_rows, num_cols, zebra=True):
    for row_idx in range(2, num_rows + 2):
        for col_idx in range(1, num_cols + 1):
            cell = sheet.cell(row=row_idx, column=col_idx)
            cell.font = BODY_FONT
            cell.alignment = WRAP_TOP
            cell.border = BORDER
            if zebra and row_idx % 2 == 0:
                cell.fill = ZEBRA_FILL


def build_features_workbook(path):
    wb = Workbook()

    # Sheet 1: How to read
    tips = wb.active
    tips.title = 'How to Read'
    tips_lines = [
        ('Kirana Mart - Features Inventory', True, 16),
        ('', False, 11),
        ('What is this sheet?', True, 12),
        ('A catalog of every user-facing feature in the Kirana Mart app as of today. Use this to learn what the app can do, who each feature is for, and where to find it.', False, 11),
        ('', False, 11),
        ('How to use it', True, 12),
        ('1. Filter the Features tab by Role to see only what applies to your testing role (Customer / Shop Owner / Delivery Partner / Admin).', False, 11),
        ('2. Filter by Status to focus on Live features (skip Deferred ones for now).', False, 11),
        ('3. Each feature has a "How to Access" cell telling you exactly where to find it in the app.', False, 11),
        ('4. The "PR Reference" column links the feature to the development pull request - useful when reporting bugs (mention the PR number).', False, 11),
        ('', False, 11),
        ('Status colors', True, 12),
        ('Green (Live) - feature is working and ready to test.', False, 11),
        ('Yellow (In Progress) - feature is being built; may be partially testable.', False, 11),
        ('Red (Deferred) - feature is paused or has a known bug. See Notes.', False, 11),
        ('', False, 11),
        ('Updates', True, 12),
        ('This sheet will be updated as new PRs ship. The Excel file in Google Drive is the single source of truth - re-download or refresh after a deploy.', False, 11),
        ('', False, 11),
        ('Roles in this app', True, 12),
        ('Customer - buys groceries via the app.', False, 11),
        ('Shop Owner - runs a kirana store; manages menu, accepts orders, marks them ready.', False, 11),
        ('Delivery Partner - picks up ready orders and delivers them.', False, 11),
        ('Admin - oversees the platform; approves shops/partners, manages users, handles refunds.', False, 11),
        ('All Users - features anyone can use regardless of role.', False, 11),
    ]
    for i, (text, bold, size) in enumerate(tips_lines, start=1):
        cell = tips.cell(row=i, column=1, value=text)
        cell.font = Font(name='Arial', size=size, bold=bold)
        cell.alignment = Alignment(wrap_text=True, vertical='top')
    tips.column_dimensions['A'].width = 100
    tips.sheet_view.showGridLines = False

    # Sheet 2: Features
    sheet = wb.create_sheet('Features')
    headers = [
        'Feature ID', 'Category', 'Role', 'Feature Name', 'Description',
        'How to Access', 'Status', 'PR Reference', 'Notes',
    ]
    sheet.append(headers)
    for row in FEATURES:
        sheet.append(list(row))

    style_header_row(sheet, len(headers))
    style_body_cells(sheet, len(FEATURES), len(headers))

    widths = [12, 18, 18, 30, 50, 35, 12, 14, 35]
    for i, w in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(i)].width = w

    # Status color coding (col 7)
    for row_idx in range(2, len(FEATURES) + 2):
        status_cell = sheet.cell(row=row_idx, column=7)
        if status_cell.value == 'Live':
            status_cell.fill = PatternFill('solid', start_color=LIGHT_GREEN)
            status_cell.font = Font(name='Arial', size=10, bold=True, color='065F46')
        elif status_cell.value == 'Deferred':
            status_cell.fill = PatternFill('solid', start_color='FEE2E2')
            status_cell.font = Font(name='Arial', size=10, bold=True, color='991B1B')
        elif status_cell.value == 'In Progress':
            status_cell.fill = PatternFill('solid', start_color=LIGHT_YELLOW)
            status_cell.font = Font(name='Arial', size=10, bold=True, color='92400E')

    wb.save(path)


def build_test_cases_workbook(path):
    wb = Workbook()

    # Sheet 1: How to use
    intro = wb.active
    intro.title = 'How to Use'
    intro_lines = [
        ('Kirana Mart - Test Cases', True, 16),
        ('', False, 11),
        ('What is this sheet?', True, 12),
        ('Every test case for the app, organized so you can systematically verify each feature works correctly. The goal: if every test passes, the app is solid.', False, 11),
        ('', False, 11),
        ('How to test', True, 12),
        ("1. Go to the 'Test Cases' tab.", False, 11),
        ('2. Filter by Role to see only your tests (Customer / Shop Owner / Delivery / Admin / All).', False, 11),
        ('3. For each test case, read Pre-conditions and ensure the app is in that state.', False, 11),
        ('4. Follow Test Steps exactly. Use the Test Data column if needed.', False, 11),
        ('5. Compare what you see to Expected Result.', False, 11),
        ('6. Fill in the BLANK columns:', False, 11),
        ('   - Actual Result - what you actually saw', False, 11),
        ('   - Status - Pass / Fail / Blocked', False, 11),
        ('   - Severity (if Fail) - Critical / Major / Minor / Cosmetic', False, 11),
        ('   - Tested By - your name', False, 11),
        ('   - Tested Date - today', False, 11),
        ('   - Comments - any observation, screenshot reference, bug repro hints', False, 11),
        ('', False, 11),
        ('Status definitions', True, 12),
        ('Pass - actual result matches expected. App works as designed.', False, 11),
        ('Fail - actual result is wrong, missing, or shows an error.', False, 11),
        ('Blocked - could not complete this test (e.g. a pre-requisite feature is broken).', False, 11),
        ('', False, 11),
        ('Severity definitions (when Fail)', True, 12),
        ('Critical - app crashes, data lost, payment fails, or test cannot continue.', False, 11),
        ('Major - feature is broken but app keeps working. User cannot complete a key flow.', False, 11),
        ('Minor - feature works but has a noticeable issue (wrong text, slow response, ugly layout).', False, 11),
        ("Cosmetic - only visual or wording. Doesn't affect function.", False, 11),
        ('', False, 11),
        ('Test Case ID - Feature ID mapping', True, 12),
        ('Each test case references a Feature ID (e.g. FE-042). Look that up in the Features Inventory sheet to understand what the feature does.', False, 11),
        ('', False, 11),
        ('Reporting bugs', True, 12),
        ('When marking a test Fail, please include in Comments:', False, 11),
        ('   - Exact device and OS version (e.g. iPhone 13, iOS 17.4)', False, 11),
        ('   - Whether it is reproducible every time or intermittent', False, 11),
        ('   - Any error messages on screen (verbatim)', False, 11),
        ('   - A screenshot if possible (attach in the Google Drive folder)', False, 11),
        ('', False, 11),
        ('Known issues (skip these for now)', True, 12),
        ('- Menu image upload (FE-073, TC-113) - known broken, deferred fix.', False, 11),
        ('- Web app preview (FE-204) - paused; do not test on web.', False, 11),
    ]
    for i, (text, bold, size) in enumerate(intro_lines, start=1):
        cell = intro.cell(row=i, column=1, value=text)
        cell.font = Font(name='Arial', size=size, bold=bold)
        cell.alignment = Alignment(wrap_text=True, vertical='top')
    intro.column_dimensions['A'].width = 110
    intro.sheet_view.showGridLines = False

    # Sheet 2: Test Cases
    sheet = wb.create_sheet('Test Cases')
    headers = [
        'TC ID', 'Feature ID', 'Role', 'Test Case Title', 'Type', 'Priority',
        'Pre-conditions', 'Test Steps', 'Test Data', 'Expected Result',
        'Actual Result', 'Status', 'Severity', 'Tested By', 'Tested Date', 'Comments',
    ]
    sheet.append(headers)
    for row in TEST_CASES:
        sheet.append(list(row) + ['', '', '', '', '', ''])

    style_header_row(sheet, len(headers))
    style_body_cells(sheet, len(TEST_CASES), len(headers))

    widths = [10, 12, 14, 35, 12, 11, 35, 50, 22, 45, 35, 12, 12, 18, 14, 30]
    for i, w in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(i)].width = w

    # Color-code Type (col 5)
    for row_idx in range(2, len(TEST_CASES) + 2):
        type_cell = sheet.cell(row=row_idx, column=5)
        if type_cell.value == 'Positive':
            type_cell.fill = PatternFill('solid', start_color=LIGHT_GREEN)
            type_cell.font = Font(name='Arial', size=10, bold=True, color='065F46')
        elif type_cell.value == 'Negative':
            type_cell.fill = PatternFill('solid', start_color='FEE2E2')
            type_cell.font = Font(name='Arial', size=10, bold=True, color='991B1B')
        elif type_cell.value == 'Edge case':
            type_cell.fill = PatternFill('solid', start_color=LIGHT_YELLOW)
            type_cell.font = Font(name='Arial', size=10, bold=True, color='92400E')

        # Priority text color
        priority_cell = sheet.cell(row=row_idx, column=6)
        if priority_cell.value == 'Critical':
            priority_cell.font = Font(name='Arial', size=10, bold=True, color='991B1B')
        elif priority_cell.value == 'High':
            priority_cell.font = Font(name='Arial', size=10, bold=True, color='9A3412')
        elif priority_cell.value == 'Medium':
            priority_cell.font = Font(name='Arial', size=10, bold=True, color='92400E')
        elif priority_cell.value == 'Low':
            priority_cell.font = Font(name='Arial', size=10, color='4B5563')

    # Tester-input columns get a subtle tint
    tester_input_fill = PatternFill('solid', start_color='FFFBEB')
    for row_idx in range(2, len(TEST_CASES) + 2):
        for col_idx in [11, 12, 13, 14, 15, 16]:
            sheet.cell(row=row_idx, column=col_idx).fill = tester_input_fill

    # Dropdowns for Status + Severity
    status_dv = DataValidation(
        type='list',
        formula1='"Pass,Fail,Blocked,Not Tested"',
        allow_blank=True,
    )
    status_dv.error = 'Select Pass, Fail, Blocked, or Not Tested.'
    status_dv.errorTitle = 'Invalid Status'
    sheet.add_data_validation(status_dv)
    status_dv.add(f'L2:L{len(TEST_CASES) + 1}')

    severity_dv = DataValidation(
        type='list',
        formula1='"Critical,Major,Minor,Cosmetic"',
        allow_blank=True,
    )
    severity_dv.error = 'Select Critical, Major, Minor, or Cosmetic.'
    severity_dv.errorTitle = 'Invalid Severity'
    sheet.add_data_validation(severity_dv)
    severity_dv.add(f'M2:M{len(TEST_CASES) + 1}')

    # Sheet 3: Summary
    summary = wb.create_sheet('Summary')
    summary['A1'] = 'Kirana Mart - Testing Summary'
    summary['A1'].font = Font(name='Arial', size=16, bold=True)

    summary['A3'] = 'Total test cases'
    summary['B3'] = len(TEST_CASES)
    summary['A4'] = 'Tested (Pass + Fail + Blocked)'
    summary['B4'] = (
        f"=COUNTIF('Test Cases'!L2:L{len(TEST_CASES) + 1},\"Pass\")"
        f"+COUNTIF('Test Cases'!L2:L{len(TEST_CASES) + 1},\"Fail\")"
        f"+COUNTIF('Test Cases'!L2:L{len(TEST_CASES) + 1},\"Blocked\")"
    )
    summary['A5'] = 'Passed'
    summary['B5'] = f"=COUNTIF('Test Cases'!L2:L{len(TEST_CASES) + 1},\"Pass\")"
    summary['A6'] = 'Failed'
    summary['B6'] = f"=COUNTIF('Test Cases'!L2:L{len(TEST_CASES) + 1},\"Fail\")"
    summary['A7'] = 'Blocked'
    summary['B7'] = f"=COUNTIF('Test Cases'!L2:L{len(TEST_CASES) + 1},\"Blocked\")"
    summary['A8'] = 'Not yet tested'
    summary['B8'] = "=B3-B4"
    summary['A9'] = 'Pass rate'
    summary['B9'] = "=IF(B4=0,0,B5/B4)"
    summary['B9'].number_format = '0.0%'

    summary['A11'] = 'Failures by severity'
    summary['A11'].font = Font(name='Arial', size=12, bold=True)
    summary['A12'] = 'Critical'
    summary['B12'] = (
        f"=COUNTIFS('Test Cases'!L2:L{len(TEST_CASES) + 1},\"Fail\","
        f"'Test Cases'!M2:M{len(TEST_CASES) + 1},\"Critical\")"
    )
    summary['A13'] = 'Major'
    summary['B13'] = (
        f"=COUNTIFS('Test Cases'!L2:L{len(TEST_CASES) + 1},\"Fail\","
        f"'Test Cases'!M2:M{len(TEST_CASES) + 1},\"Major\")"
    )
    summary['A14'] = 'Minor'
    summary['B14'] = (
        f"=COUNTIFS('Test Cases'!L2:L{len(TEST_CASES) + 1},\"Fail\","
        f"'Test Cases'!M2:M{len(TEST_CASES) + 1},\"Minor\")"
    )
    summary['A15'] = 'Cosmetic'
    summary['B15'] = (
        f"=COUNTIFS('Test Cases'!L2:L{len(TEST_CASES) + 1},\"Fail\","
        f"'Test Cases'!M2:M{len(TEST_CASES) + 1},\"Cosmetic\")"
    )

    summary['A17'] = 'Test cases by role'
    summary['A17'].font = Font(name='Arial', size=12, bold=True)
    roles = ['All', 'Customer', 'Shop Owner', 'Delivery Partner', 'Admin']
    for i, role in enumerate(roles, start=18):
        summary[f'A{i}'] = role
        summary[f'B{i}'] = f"=COUNTIF('Test Cases'!C2:C{len(TEST_CASES) + 1},\"{role}\")"

    summary.column_dimensions['A'].width = 35
    summary.column_dimensions['B'].width = 15

    wb.save(path)


if __name__ == '__main__':
    features_path = os.path.join(OUTPUT_DIR, 'kirana-mart-features.xlsx')
    test_cases_path = os.path.join(OUTPUT_DIR, 'kirana-mart-test-cases.xlsx')

    print(f'Building features workbook: {features_path}')
    build_features_workbook(features_path)
    print(f'  {len(FEATURES)} features written.')

    print(f'Building test cases workbook: {test_cases_path}')
    build_test_cases_workbook(test_cases_path)
    print(f'  {len(TEST_CASES)} test cases written.')

    print('\nDone. Upload both .xlsx files to Google Drive.')
